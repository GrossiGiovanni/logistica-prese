# Converte i 5 file Excel del cliente in un unico bundle JSON per l'import Prisma.
# Esegue il parsing semi-strutturato del campo NOTE delle prese
# (PLT / MTL / MC / KG / ORE / MOTRICE / destinazione finale).
#
# Esecuzione: py -3 prisma/seed-data/build_bundle.py
# Output:     prisma/seed-data/bundle.json
import json
import re
from pathlib import Path
import openpyxl

DOWNLOADS = Path(r"C:\Users\giova\Downloads")
OUT = Path(__file__).parent / "bundle.json"

F_AUTISTI = DOWNLOADS / "AUTISTI.xlsx"
F_FISSE = DOWNLOADS / "PRESE FISSE GIORNALIERE (2).xlsx"
F_WEEK = DOWNLOADS / "PRESE WEEK 26.xlsx"
F_2306 = DOWNLOADS / "prese 23.06.xlsx"


def norm(v):
    """Trim + collassa spazi multipli; None/'' -> None."""
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s or None


def num(s):
    """'4,5' / '4.5' -> 4.5 ; None se non numerico."""
    if s is None:
        return None
    m = re.search(r"\d+(?:[.,]\d+)?", str(s))
    if not m:
        return None
    return float(m.group(0).replace(",", "."))


def parse_note(note):
    """Estrae quantità, fascia oraria, mezzo e destinazione dal testo NOTE."""
    out = {
        "pallets": None,
        "loadingMeters": None,
        "volumeM3": None,
        "weightKg": None,
        "timeWindow": "ANYTIME",
        "timeFrom": None,
        "requiresMotrice": False,
        "destination": None,
    }
    if not note:
        return out
    up = note.upper()

    # Pallet: somma di tutte le occorrenze "<n> PLT"
    plts = re.findall(r"(\d+)\s*PLT", up)
    if plts:
        out["pallets"] = sum(int(x) for x in plts)

    # Metri lineari, metri cubi, peso
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*MTL", up)
    if m:
        out["loadingMeters"] = float(m.group(1).replace(",", "."))
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*MC", up)
    if m:
        out["volumeM3"] = float(m.group(1).replace(",", "."))
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*KG", up)
    if m:
        out["weightKg"] = float(m.group(1).replace(",", "."))

    # Orario: "ORE 14", "ORE 15.3" -> HH:MM, fascia SPECIFIC
    m = re.search(r"ORE\s*(\d{1,2})(?:[.,:](\d{1,2}))?", up)
    if m:
        hh = int(m.group(1))
        mm = int(m.group(2)) if m.group(2) else 0
        if mm < 10 and m.group(2) and len(m.group(2)) == 1:
            mm = mm * 10  # "15.3" -> 15:30
        out["timeFrom"] = f"{hh:02d}:{mm:02d}"
        out["timeWindow"] = "SPECIFIC"
    elif "POMERIGG" in up:
        out["timeWindow"] = "AFTERNOON"
    elif "MATTIN" in up:
        out["timeWindow"] = "MORNING"

    # Mezzo richiesto
    if "MOTRICE" in up:
        out["requiresMotrice"] = True

    # Destinazione finale: testo dopo " X " (es. "19 PLT X MD GRICIGNANO")
    m = re.search(r"\bX\s+([A-Z0-9][A-Z0-9 /.'\-]+)$", up)
    if m:
        out["destination"] = norm(m.group(1))

    return out


def read_autisti():
    wb = openpyxl.load_workbook(F_AUTISTI, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    drivers = []
    for r in rows[1:]:
        codice, autista, mezzo, targa, titolare = (r + (None,) * 5)[:5]
        if not norm(autista):
            continue
        drivers.append({
            "codice": norm(codice),
            "autista": norm(autista),
            "mezzo": (norm(mezzo) or "BILICO").upper(),  # BILICO / MOTRICE
            "targa": norm(targa),
            "titolare": norm(titolare),
        })
    return drivers


def read_fisse():
    wb = openpyxl.load_workbook(F_FISSE, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    fixed = []
    for r in rows:
        mittente, street, city, prov = (tuple(r) + (None,) * 4)[:4]
        m, s, c = norm(mittente), norm(street), norm(city)
        # Salta titolo e intestazione
        if not m or m.upper() in ("MITTENTE", "PRESE FISSE GIORNALIERE"):
            continue
        if not s or not c:
            continue
        fixed.append({"mittente": m, "street": s, "city": c, "province": norm(prov)})
    return fixed


def read_prese(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[1:]:
        cells = (tuple(r) + (None,) * 10)[:10]
        numero, data, mittente, indirizzo, localita, prov, note, colli, peso, mc = cells
        m = norm(mittente)
        if not m:
            continue
        date_str = None
        if data is not None:
            date_str = str(data)[:10]  # datetime -> 'YYYY-MM-DD'
        note_s = norm(note)
        parsed = parse_note(note_s)
        # Colonne numeriche esplicite hanno priorità se presenti
        if num(colli) is not None:
            parsed["colli"] = int(num(colli))
        else:
            parsed["colli"] = None
        if num(peso) is not None:
            parsed["weightKg"] = num(peso)
        if num(mc) is not None:
            parsed["volumeM3"] = num(mc)

        out.append({
            "numero": norm(numero),
            "date": date_str,
            "mittente": m,
            "street": norm(indirizzo),
            "city": norm(localita),
            "province": norm(prov),
            "note": note_s,
            **parsed,
        })
    return out


def main():
    drivers = read_autisti()
    fixed = read_fisse()

    pickups = read_prese(F_WEEK)
    seen = {p["numero"] for p in pickups if p["numero"]}
    extra = 0
    for p in read_prese(F_2306):
        if p["numero"] and p["numero"] in seen:
            continue
        if p["numero"]:
            seen.add(p["numero"])
        pickups.append(p)
        extra += 1

    bundle = {"drivers": drivers, "fixed": fixed, "pickups": pickups}
    OUT.write_text(json.dumps(bundle, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Scritto {OUT}")
    print(f"  autisti/mezzi: {len(drivers)}")
    print(f"  prese fisse:   {len(fixed)}")
    print(f"  prese:         {len(pickups)} (di cui {extra} extra da prese 23.06)")
    # Riepilogo parsing
    with_plt = sum(1 for p in pickups if p["pallets"] is not None)
    with_time = sum(1 for p in pickups if p["timeWindow"] != "ANYTIME")
    with_dest = sum(1 for p in pickups if p["destination"])
    print(f"  con pallet:    {with_plt}")
    print(f"  con orario:    {with_time}")
    print(f"  con destinaz.: {with_dest}")


if __name__ == "__main__":
    main()
